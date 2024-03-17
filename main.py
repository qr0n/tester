from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pyautogui
import pygetwindow as gw
import time
import subprocess
import json
import re

with open("F:/projects/compsci/tester/config.json", "r") as E:
    config = json.load(E)

c_file = config["c_file"]

screenshot_path = config["screenshot_path"]
left, top, width, height = config["left"], config["top"], config["width"], config["height"]

alphabet_dict = {i: chr(i + ord('A') - 1) for i in range(1, 27)}

workbook_path = config["workbook_path"]
workbook = load_workbook(workbook_path)
sheet = workbook['Sheet1']

# Function extractor (Just like extractor.py but in one file, might compile to executable later.)

def extract_function_signatures(file_path):
    with open(file_path, 'r') as file:
        content = file.read()

    # Regular expression to find all function signatures
    pattern = re.compile(r'\b(?:\w+\s+)?(\w+\s+\w+\([^)]*\))\s*{')
    matches = pattern.finditer(content)

    # Extract function signatures
    function_signatures = [match.group(1) for match in matches]

    return function_signatures

def extract(c_file_path):
    all_function_signatures = extract_function_signatures(c_file_path)

    if all_function_signatures:
        print("All Function Signatures Found:")
        for i, function_signature in enumerate(all_function_signatures, start=1):
            print(f"Function {i}: {function_signature}")
    else:
        print("No functions found in the provided C file.")

# End extractor.py

class ExcelFileManagement:
    @staticmethod
    def add_value(cell_id, text):
        try:
            sheet[cell_id] = text
            workbook.save(workbook_path)
            return "Done"
        except Exception as E:
                return E
        
    @staticmethod
    def add_image_to_cell(cell_id, image_path, width, height):
        try:
            img = Image(image_path)
            img.width = width
            img.height = height
            sheet.add_image(img, cell_id)
            workbook.save(workbook_path)
            return "Done"
        except Exception as E:
            return E
            
    @staticmethod
    def populate_table():
        headers = ["Test number", "Test Type", "Variables", "Module", "Input", "Expected result", "Actual result", "Screenshot"]
    # Calculate the starting cell for populating data
        if sheet.max_row == 1:
            start_row = 1
        else:
            start_row = sheet.max_row + 5
        start_cell = f"{alphabet_dict[1]}{start_row}"

        for i, header in enumerate(headers, start=1):
            sheet.cell(row=i + start_row - 1, column=1, value=header)

        print(f"Starting from cell: {start_cell}")
        workbook.save(workbook_path)
        
    @staticmethod
    def add_results(test_type, variables, module, input_data, expected):
        ExcelFileManagement.populate_table()
        try:
            print("compiling")
            process = subprocess.Popen(["gcc", c_file, "-o", "output.exe"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            _, error = process.communicate()
            if error:
                raise Exception(f"Compilation Error: {error.decode('utf-8')}")
            print("finished compiling")

            print("executing")
            process = subprocess.Popen(["output.exe"], stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            output, error = process.communicate(input=input_data.encode('utf-8'))
            if error:
                raise Exception(f"Execution Error: {error.decode('utf-8')}")
        # Find the next available row
            if sheet.max_row == 1:
                row = sheet.max_row
            else:
                row = sheet.max_row - 7
        
        # Define the starting column index
            start_col = 2  # Column B
        
        # Set values in the corresponding cells
            sheet.cell(row=row + 1, column=start_col).value = test_type
            sheet.cell(row=row + 2, column=start_col).value = variables
            
            sheet.cell(row=row + 3, column=start_col).value = module
            sheet.cell(row=row + 4, column=start_col).value = input_data
            
            sheet.cell(row=row + 5, column=start_col).value = expected
            sheet.cell(row=row + 6, column=start_col).value = output.decode("utf-8")

            ss_path = f"{screenshot_path}{row + 7}.png"
            ScreenshotManagement.take_screenshot(ss_path)

            ExcelFileManagement.add_image_to_cell(cell_id=f"B{row + 7}", image_path=ss_path, width=100, height=100)

        # Save the workbook
            workbook.save(workbook_path)
            workbook.close()
            return "Results added successfully"
        except Exception as e:
            return str(e)

class ScreenshotManagement:
    class Helper:
        def focus_window_by_title(window_title):
            try:
                window = gw.getWindowsWithTitle(window_title)[0]
                window.activate()
                return True
            except IndexError:
                print(f"Window with title '{window_title}' not found.")
                return False
            
    @staticmethod
    def take_screenshot(path_to_save):
        ScreenshotManagement.Helper.focus_window_by_title("Windows PowerShell")
        time.sleep(1)
        screenshot = pyautogui.screenshot(region=(left, top, width, height))
        screenshot.save(path_to_save)

class UserManagement:
    @staticmethod
    def prompt():
        type_input = input("What type of test are you running? (Normal/Erroneous/Extreme/Incomplete)?\n> ")
        variable_input = input("Paste the variables your code is using here\n> ")

        print(f"Reading file located at {c_file}")
        print("Extracting function signatures...")
        extract(c_file)
        module_input = input("Select the function you're testing from the list above (or if it is not there enter a new one)\n> ")

        input_input = input("Enter the input data:\n> ")
        expected_input = input("What do you expect this code to return?\n> ")

        result = ExcelFileManagement.add_results(test_type=type_input, variables=variable_input, module=module_input, input_data=input_input, expected=expected_input)
        if result == "Results added successfully":
            try:
                UserManagement.Helper.prompt()
            except KeyboardInterrupt as E:
                print("Saving workbook...")
                workbook.save(workbook_path)
                workbook.close()
                print("Closing file...\nExitting.")
                exit()
        else:
            print(f"Error: {result}")

UserManagement.prompt()
